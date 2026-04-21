
import argparse
import io
import math
import os
import re
import sys
import time
import uuid
from collections import Counter

from dotenv import load_dotenv
from openpyxl import load_workbook

try:
    from supabase import create_client
except Exception:
    create_client = None

try:
    if hasattr(sys.stdout, "buffer"):
        sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
    if hasattr(sys.stderr, "buffer"):
        sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8", errors="replace")
except Exception:
    pass


DEFAULT_EXCEL_PATH = r"E:\OneDrive - Salasar Services Pvt. Ltd\Desktop\Baja Motor Extended Warranty  Retail Pricelist_2026.xlsx"


def log_info(message):
    print(f"[INFO] {message}")


def log_ok(message):
    print(f"[OK] {message}")


def log_warn(message):
    print(f"[WARN] {message}")


def log_error(message):
    print(f"[ERROR] {message}")


def render_progress(prefix, current, total, width=30):
    total = max(int(total), 1)
    current = max(0, min(int(current), total))
    filled = int(width * current / total)
    bar = "#" * filled + "-" * (width - filled)
    percent = (current / total) * 100
    print(f"\r{prefix} [{bar}] {current}/{total} ({percent:5.1f}%)", end="", flush=True)


def finish_progress():
    print()



def clean_text(val):
    if val is None:
        return ""
    s = str(val).replace("\xa0", " ").replace("\n", " ").strip()
    return re.sub(r"\s+", " ", s)


def is_blank(val):
    return clean_text(val) == ""


def to_price(val):
    if val is None:
        return None
    if isinstance(val, (int, float)) and not isinstance(val, bool):
        try:
            if math.isnan(val):
                return None
        except Exception:
            pass
        return float(val)

    s = clean_text(val).replace(",", "").replace(" ", "")
    if not s or s.upper() in {"NA", "N/A", "-", "NONE"}:
        return None

    try:
        return float(s)
    except ValueError:
        return None


def norm_fuel(raw):
    s = clean_text(raw).upper()
    mapping = {
        "P": "Petrol",
        "PETROL": "Petrol",
        "GASOLINE": "Petrol",
        "D": "Diesel",
        "DIESEL": "Diesel",
        "B": "EV",
        "BEV": "EV",
        "EV": "EV",
        "CNG": "CNG",
        "PH": "Hybrid",
        "HYBRID": "Hybrid",
        "P/D": "Petrol/Diesel",
        "DIESEL & PETROL": "Petrol/Diesel",
        "PETROL/DIESEL": "Petrol/Diesel",
        "NA": "N/A",
        "N/A": "N/A",
        "": "N/A",
    }
    return mapping.get(s, clean_text(raw) or "N/A")


def norm_trans(raw):
    s = clean_text(raw).upper()
    mapping = {
        "MT": "Manual",
        "M/T": "Manual",
        "MANUAL": "Manual",
        "AT": "Automatic",
        "A/T": "Automatic",
        "AUTOMATIC": "Automatic",
        "AMT": "AMT",
        "CVT": "CVT",
        "DCA": "DCA",
        "NA": "N/A",
        "N/A": "N/A",
        "ALL VARIANTS": "N/A",
        "MT/AT": "N/A",
        "": "N/A",
    }
    return mapping.get(s, clean_text(raw) or "N/A")


def normalize_brand(name):
    s = clean_text(name)
    mapping = {
        "VW": "Volkswagen",
        "JEEP": "Jeep",
        "MBI": "Mercedes-Benz",
        "RANGE ROVER": "Land Rover",
    }
    return mapping.get(s.upper(), s)


def parse_kms(raw):
    s = clean_text(raw).upper()
    if not s:
        return None
    if "UNL" in s or "UNLIMIT" in s:
        return None

    s = s.replace(",", "")
    m = re.search(r"(\d+(?:\.\d+)?)\s*L", s)
    if m:
        return int(float(m.group(1)) * 100000)

    digits = re.findall(r"\d+", s)
    if not digits:
        return None

    if len(digits) == 1:
        return int(digits[0])

    return int("".join(digits))


def infer_fuel_trans(text):
    s = clean_text(text).upper()
    fuel = "N/A"
    trans = "N/A"

    if "CNG" in s:
        fuel = "CNG"
    elif "DIESEL" in s or re.search(r"\bD\b", s):
        fuel = "Diesel"
    elif "PETROL" in s or re.search(r"\bP\b", s):
        fuel = "Petrol"
    elif "EV" in s or "BEV" in s:
        fuel = "EV"
    elif "HEV" in s or "HYBRID" in s or re.search(r"\bPH\b", s):
        fuel = "Hybrid"

    if "AMT" in s:
        trans = "AMT"
    elif "CVT" in s:
        trans = "CVT"
    elif "DCA" in s:
        trans = "DCA"
    elif "A/T" in s or re.search(r"\bAT\b", s) or "AUTOMATIC" in s:
        trans = "Automatic"
    elif "M/T" in s or re.search(r"\bMT\b", s) or "MANUAL" in s:
        trans = "Manual"

    return fuel, trans


def extract_duration_months(plan_name, oem_months):
    s = clean_text(plan_name)

    m = re.search(r"Additional\s+(\d+)\s+year", s, re.I)
    if m:
        return int(m.group(1)) * 12

    m = re.match(r"(\d+)\+(\d+)", s)
    if m:
        return int(m.group(2)) * 12

    m = re.search(r"(\d+)\s*Months", s, re.I)
    if m:
        return int(m.group(1))

    years = [int(x) for x in re.findall(r"(\d+)(?:st|nd|rd|th)?", s)]
    if years:
        max_year = max(years)
        return max(12, max_year * 12 - int(oem_months))

    return None


def month_range_to_days(header_text):
    text = clean_text(header_text).lower()
    nums = [int(x) for x in re.findall(r"\d+", text)]
    if len(nums) >= 2:
        low = 0 if nums[0] == 0 else nums[0] * 30 + 1
        high = (nums[1] + 1) * 30
        return low, high
    return 0, 1095


def build_record(brand, model, variant, fuel, transmission, oem_months, oem_kms, plan_name, max_kms, day_ranges, prices):
    tiers = []
    for (min_days, max_days), price in zip(day_ranges, prices):
        value = to_price(price)
        if value is None:
            continue
        tiers.append(
            {
                "min_days": int(min_days),
                "max_days": int(max_days),
                "price_inr": float(value),
                "is_active": True,
            }
        )

    if not tiers:
        return None

    return {
        "brand": normalize_brand(brand),
        "model": clean_text(model),
        "variant": clean_text(variant),
        "fuel": norm_fuel(fuel),
        "transmission": norm_trans(transmission),
        "oem_warranty_months": int(oem_months) if oem_months is not None else None,
        "oem_warranty_kms": int(oem_kms) if oem_kms is not None else None,
        "plan_name": clean_text(plan_name),
        "plan_code": None,
        "duration_months": extract_duration_months(plan_name, oem_months or 0),
        "max_kms": int(max_kms) if max_kms is not None else None,
        "tiers": tiers,
    }


def parse_kia(workbook):
    ws = workbook["Kia"]
    plan_names = [f"{clean_text(ws.cell(1, c).value)} {clean_text(ws.cell(2, c).value)}".strip() for c in range(3, 7)]
    age_map = {
        "0-3 Months": (0, 90),
        "4-12 Months": (91, 365),
        "13-24 Months": (366, 730),
        "25-34 Months": (731, 1020),
    }

    current_models = []
    for row in range(3, ws.max_row + 1):
        model_cell = ws.cell(row, 1).value
        age_text = clean_text(ws.cell(row, 2).value)
        if model_cell:
            current_models = []
            for chunk in [clean_text(x) for x in str(model_cell).split("/") if clean_text(x)]:
                m = re.match(r"(.+?)-\s*(Petrol|Diesel)$", chunk, re.I)
                if m:
                    current_models.append((clean_text(m.group(1)), norm_fuel(m.group(2))))
                else:
                    current_models.append((chunk, "N/A"))

        if age_text not in age_map:
            continue

        for model, fuel in current_models:
            variant = f"{model} - {fuel}" if fuel != "N/A" else model
            for idx, plan_name in enumerate(plan_names):
                record = build_record(
                    "Kia",
                    model,
                    variant,
                    fuel,
                    "N/A",
                    36,
                    None,
                    plan_name,
                    parse_kms(plan_name),
                    [age_map[age_text]],
                    [ws.cell(row, 3 + idx).value],
                )
                if record:
                    yield record


def parse_tata(workbook):
    ws = workbook["Tata"]
    current_model = None
    current_plan = None
    day_ranges = [(0, 90), (91, 180), (181, 1095)]

    for row in range(1, ws.max_row + 1):
        col_a = clean_text(ws.cell(row, 1).value)
        col_b = clean_text(ws.cell(row, 2).value)

        if col_a and "DAYS FROM DATE OF SALE" in col_b.upper():
            current_model = col_a.title()
            current_plan = None
            continue

        if col_a.upper().startswith("ADDITIONAL"):
            current_plan = col_a
            continue

        if current_model and current_plan and col_a and to_price(ws.cell(row, 2).value) is not None:
            fuel, transmission = infer_fuel_trans(f"{col_a} {current_model}")
            record = build_record(
                "Tata",
                current_model,
                col_a,
                fuel,
                transmission,
                36,
                None,
                current_plan,
                parse_kms(current_plan),
                day_ranges,
                [ws.cell(row, 2).value, ws.cell(row, 3).value, ws.cell(row, 4).value],
            )
            if record:
                yield record


def parse_hyundai(workbook):
    ws = workbook["Hyundai"]
    current_model = None
    plans = [
        (clean_text(ws.cell(1, 3).value), 3),
        (clean_text(ws.cell(1, 6).value), 6),
        (clean_text(ws.cell(1, 9).value), 9),
    ]
    day_ranges = [(0, 90), (91, 365), (366, 1095)]

    for row in range(3, ws.max_row + 1):
        if clean_text(ws.cell(row, 1).value):
            current_model = clean_text(ws.cell(row, 1).value)

        fuel_trans = clean_text(ws.cell(row, 2).value)
        if not current_model or not fuel_trans:
            continue

        parts = fuel_trans.split("-")
        fuel = norm_fuel(parts[0])
        transmission = norm_trans(parts[1] if len(parts) > 1 else fuel_trans)
        variant = f"{current_model} {fuel_trans}"

        for plan_name, start_col in plans:
            record = build_record(
                "Hyundai",
                current_model,
                variant,
                fuel,
                transmission,
                36,
                None,
                plan_name,
                parse_kms(plan_name),
                day_ranges,
                [ws.cell(row, start_col + i).value for i in range(3)],
            )
            if record:
                yield record


def parse_renault(workbook):
    ws = workbook["Renault"]
    current_fuel = "N/A"
    current_model = None
    day_ranges = [(0, 90), (91, 365), (366, 730), (731, 1095)]

    for row in range(3, ws.max_row + 1):
        col_a = clean_text(ws.cell(row, 1).value)
        col_b = clean_text(ws.cell(row, 2).value)

        if col_a.upper().startswith("MODEL"):
            m = re.search(r"\(([^)]+)\)", col_a)
            current_fuel = norm_fuel(m.group(1) if m else "N/A")
            current_model = None
            continue

        if col_a:
            current_model = col_a

        if current_model and col_b:
            record = build_record(
                "Renault",
                current_model,
                current_model,
                current_fuel,
                "N/A",
                36,
                None,
                col_b,
                parse_kms(col_b),
                day_ranges,
                [ws.cell(row, c).value for c in range(3, 7)],
            )
            if record:
                yield record


def parse_maruti(workbook):
    ws = workbook["Maruti New"]
    day_ranges = [(0, 30), (31, 365), (366, 730), (731, 1095)]
    plans = [("12 Months/20k", 8), ("24 Months/40k", 12), ("36 Months/60k", 16)]

    for row in range(4, ws.max_row + 1):
        variant = clean_text(ws.cell(row, 1).value)
        model = clean_text(ws.cell(row, 3).value)
        if not variant or not model:
            continue

        fuel = norm_fuel(ws.cell(row, 4).value)
        transmission = norm_trans(ws.cell(row, 5).value)
        oem_months = int(ws.cell(row, 6).value)
        oem_kms = parse_kms(ws.cell(row, 7).value)

        for plan_name, start_col in plans:
            record = build_record(
                "Maruti",
                model,
                variant,
                fuel,
                transmission,
                oem_months,
                oem_kms,
                plan_name,
                parse_kms(plan_name),
                day_ranges,
                [ws.cell(row, start_col + i).value for i in range(4)],
            )
            if record:
                yield record


def parse_mahindra(workbook):
    ws = workbook["Mahindra"]
    current_model = None
    day_ranges = [(0, 60), (61, 730), (731, 1095)]
    plans = [("4th Year/120000 kms", 4), ("4th & 5th Year/150000 kms", 7)]

    for row in range(6, ws.max_row + 1):
        if clean_text(ws.cell(row, 1).value):
            current_model = clean_text(ws.cell(row, 1).value)

        fuel_raw = clean_text(ws.cell(row, 3).value)
        if not current_model or not fuel_raw:
            continue

        fuel = norm_fuel(fuel_raw)
        transmission = norm_trans(ws.cell(row, 2).value)
        variant = f"{current_model} - {fuel} - {transmission}"

        for plan_name, start_col in plans:
            record = build_record(
                "Mahindra",
                current_model,
                variant,
                fuel,
                transmission,
                36,
                None,
                plan_name,
                parse_kms(plan_name),
                day_ranges,
                [ws.cell(row, start_col + i).value for i in range(3)],
            )
            if record:
                yield record


def parse_citroen(workbook):
    ws = workbook["Citroen"]
    row = 1

    while row <= ws.max_row:
        if clean_text(ws.cell(row, 6).value).upper() == "PRICES WITH GST":
            plan_row = row + 1
            day_row = row + 2
            data_row = row + 3

            plan_groups = []
            col = 6
            while col <= ws.max_column:
                plan_name = clean_text(ws.cell(plan_row, col).value)
                if plan_name:
                    current_day_ranges = []
                    cursor = col
                    while cursor <= ws.max_column and clean_text(ws.cell(day_row, cursor).value):
                        nums = [int(x) for x in re.findall(r"\d+", clean_text(ws.cell(day_row, cursor).value))]
                        if len(nums) >= 2:
                            current_day_ranges.append((nums[0], nums[1]))
                        cursor += 1
                    plan_groups.append((plan_name, col, current_day_ranges))
                    col = cursor
                else:
                    col += 1

            cursor_row = data_row
            while cursor_row <= ws.max_row and any(ws.cell(cursor_row, c).value is not None for c in range(1, 6)):
                model = clean_text(ws.cell(cursor_row, 1).value)
                if model:
                    fuel = norm_fuel(ws.cell(cursor_row, 2).value)
                    variant = clean_text(ws.cell(cursor_row, 3).value)
                    oem_months = int(ws.cell(cursor_row, 4).value)
                    oem_kms = parse_kms(ws.cell(cursor_row, 5).value)

                    for plan_name, start_col, day_ranges in plan_groups:
                        record = build_record(
                            "Citroen",
                            model,
                            variant,
                            fuel,
                            "N/A",
                            oem_months,
                            oem_kms,
                            plan_name,
                            parse_kms(plan_name),
                            day_ranges,
                            [ws.cell(cursor_row, start_col + i).value for i in range(len(day_ranges))],
                        )
                        if record:
                            yield record
                cursor_row += 1

            row = cursor_row
        else:
            row += 1


def parse_toyota(workbook):
    ws = workbook["Toyota"]
    plans = [
        (clean_text(ws.cell(2, 7).value), 7),
        (clean_text(ws.cell(2, 11).value), 11),
        (clean_text(ws.cell(2, 15).value), 15),
    ]
    day_ranges = [(0, 60), (61, 365), (366, 730), (731, 1095)]

    for row in range(5, ws.max_row + 1):
        model = clean_text(ws.cell(row, 2).value)
        if not model or clean_text(ws.cell(row, 5).value).upper() == "MONTHS":
            continue

        fuel = norm_fuel(ws.cell(row, 3).value)
        transmission = norm_trans(ws.cell(row, 4).value)
        oem_months = int(ws.cell(row, 5).value)
        oem_kms = parse_kms(ws.cell(row, 6).value)
        variant = f"{model} - {fuel} - {transmission}"

        for plan_name, start_col in plans:
            record = build_record(
                "Toyota",
                model,
                variant,
                fuel,
                transmission,
                oem_months,
                oem_kms,
                plan_name,
                parse_kms(plan_name),
                day_ranges,
                [ws.cell(row, start_col + i).value for i in range(4)],
            )
            if record:
                yield record


def parse_mg(workbook):
    ws = workbook["MG"]
    plans = [
        ("12 Months/100k", 6),
        ("24 Months/100k", 10),
        ("12 Months/UNL kms", 14),
        ("24 Months/UNL kms", 18),
    ]
    day_ranges = [(0, 30), (31, 180), (181, 365), (366, 1095)]

    for row in range(4, ws.max_row + 1):
        model = clean_text(ws.cell(row, 2).value)
        if not model:
            continue

        fuel = norm_fuel(ws.cell(row, 3).value)
        oem_months = int(ws.cell(row, 4).value)
        oem_kms = parse_kms(ws.cell(row, 5).value)
        upper_model = model.upper()
        if "AUTOMATIC" in upper_model or "CVT" in upper_model:
            transmission = "Automatic"
        elif "MANUAL" in upper_model:
            transmission = "Manual"
        else:
            transmission = "N/A"

        for plan_name, start_col in plans:
            record = build_record(
                "MG",
                model,
                model,
                fuel,
                transmission,
                oem_months,
                oem_kms,
                plan_name,
                parse_kms(plan_name),
                day_ranges,
                [ws.cell(row, start_col + i).value for i in range(4)],
            )
            if record:
                yield record


def parse_vw(workbook):
    ws = workbook["VW"]
    plans = [("5th yr/100k kms", 9), ("5th & 6th yr/150k kms", 11)]
    day_ranges = [(0, 365), (366, 1430)]

    for row in range(6, ws.max_row + 1):
        if to_price(ws.cell(row, 9).value) is None or clean_text(ws.cell(row, 6).value).upper() == "MONTHS":
            continue

        model = clean_text(ws.cell(row, 2).value)
        variant = clean_text(ws.cell(row, 3).value) or model
        fuel = norm_fuel(ws.cell(row, 4).value)
        transmission = norm_trans(ws.cell(row, 5).value)
        oem_months = int(ws.cell(row, 6).value)
        oem_kms = parse_kms(ws.cell(row, 7).value)

        for plan_name, start_col in plans:
            record = build_record(
                "Volkswagen",
                model,
                variant,
                fuel,
                transmission,
                oem_months,
                oem_kms,
                plan_name,
                parse_kms(plan_name),
                day_ranges,
                [ws.cell(row, start_col).value, ws.cell(row, start_col + 1).value],
            )
            if record:
                yield record


def parse_skoda(workbook):
    ws = workbook["Skoda"]
    row = 1

    while row <= ws.max_row:
        if clean_text(ws.cell(row, 8).value).upper() == "PRICES WITH GST":
            plan_row = row + 1
            header_row = row + 2
            data_row = row + 3

            plan_groups = []
            col = 8
            while col <= ws.max_column:
                plan_name = clean_text(ws.cell(plan_row, col).value)
                if plan_name:
                    day_ranges = []
                    cursor = col
                    while cursor <= ws.max_column and clean_text(ws.cell(header_row, cursor).value):
                        day_ranges.append(month_range_to_days(ws.cell(header_row, cursor).value))
                        cursor += 1
                    plan_groups.append((plan_name, col, day_ranges))
                    col = cursor
                else:
                    col += 1

            cursor_row = data_row
            while cursor_row <= ws.max_row and any(ws.cell(cursor_row, c).value is not None for c in range(1, 8)):
                model = clean_text(ws.cell(cursor_row, 2).value)
                if model and clean_text(ws.cell(cursor_row, 6).value).upper() != "MONTHS":
                    variant = clean_text(ws.cell(cursor_row, 3).value) or model
                    fuel = norm_fuel(ws.cell(cursor_row, 4).value)
                    transmission = norm_trans(ws.cell(cursor_row, 5).value)
                    oem_months = int(ws.cell(cursor_row, 6).value)
                    oem_kms = parse_kms(ws.cell(cursor_row, 7).value)

                    for plan_name, start_col, day_ranges in plan_groups:
                        record = build_record(
                            "Skoda",
                            model,
                            variant,
                            fuel,
                            transmission,
                            oem_months,
                            oem_kms,
                            plan_name,
                            parse_kms(plan_name),
                            day_ranges,
                            [ws.cell(cursor_row, start_col + i).value for i in range(len(day_ranges))],
                        )
                        if record:
                            yield record
                cursor_row += 1

            row = cursor_row
        else:
            row += 1


def parse_jeep(workbook):
    ws = workbook["JEEP"]
    row = 1
    day_ranges = [(0, 60), (61, 365), (366, 730), (731, 1095)]
    plans = [("5 Year/ 100K kms", 8), ("5 Year/ 125K Kms", 12), ("5 Year/ 150K Kms", 16)]

    while row <= ws.max_row:
        if clean_text(ws.cell(row, 8).value).upper() == "PRICES WITH GST":
            cursor_row = row + 3
            current_model = None

            while cursor_row <= ws.max_row and any(ws.cell(cursor_row, c).value is not None for c in range(1, 8)):
                if clean_text(ws.cell(cursor_row, 2).value):
                    current_model = clean_text(ws.cell(cursor_row, 2).value)

                variant = clean_text(ws.cell(cursor_row, 3).value)
                if current_model and variant:
                    fuel = norm_fuel(ws.cell(cursor_row, 4).value)
                    transmission = norm_trans(ws.cell(cursor_row, 5).value)
                    oem_months = int(ws.cell(cursor_row, 6).value)
                    oem_kms = parse_kms(ws.cell(cursor_row, 7).value)

                    for plan_name, start_col in plans:
                        record = build_record(
                            "Jeep",
                            current_model,
                            variant,
                            fuel,
                            transmission,
                            oem_months,
                            oem_kms,
                            plan_name,
                            parse_kms(plan_name),
                            day_ranges,
                            [ws.cell(cursor_row, start_col + i).value for i in range(4)],
                        )
                        if record:
                            yield record
                cursor_row += 1

            row = cursor_row
        else:
            row += 1


def parse_honda(workbook):
    ws = workbook["Honda"]
    day_ranges = [(0, 60), (61, 300), (301, 540), (541, 730), (731, 910), (911, 1095)]
    plans = [(clean_text(ws.cell(2, 3).value), 3), (clean_text(ws.cell(2, 9).value), 9)]

    for row in range(4, ws.max_row + 1):
        model = clean_text(ws.cell(row, 1).value)
        variant = clean_text(ws.cell(row, 2).value)
        if not model:
            continue

        inferred_fuel, inferred_trans = infer_fuel_trans(f"{variant} {model}")
        transmission = norm_trans(variant) if inferred_trans == "N/A" else inferred_trans

        for plan_name, start_col in plans:
            record = build_record(
                "Honda",
                model,
                variant,
                inferred_fuel,
                transmission,
                36,
                None,
                plan_name,
                parse_kms(plan_name),
                day_ranges,
                [ws.cell(row, start_col + i).value for i in range(6)],
            )
            if record:
                yield record


def parse_mbi(workbook):
    ws = workbook["MBI"]
    current_model = None
    age_map = {
        "0 - 6 Months": (0, 180),
        "6 - 12 Months": (181, 365),
        "1 to 2 year": (366, 730),
        "2 to 3 year": (731, 1095),
    }
    plans = [
        clean_text(ws.cell(3, 3).value),
        clean_text(ws.cell(3, 4).value),
        clean_text(ws.cell(3, 5).value),
    ]

    for row in range(4, ws.max_row + 1):
        if clean_text(ws.cell(row, 1).value):
            current_model = clean_text(ws.cell(row, 1).value)

        age_text = clean_text(ws.cell(row, 2).value)
        if current_model and age_text in age_map:
            for offset, plan_name in enumerate(plans, start=3):
                record = build_record(
                    "Mercedes-Benz",
                    current_model,
                    current_model,
                    "N/A",
                    "N/A",
                    36,
                    None,
                    plan_name,
                    None,
                    [age_map[age_text]],
                    [ws.cell(row, offset).value],
                )
                if record:
                    yield record


def parse_audi(workbook):
    ws = workbook["Audi"]
    plans = [(clean_text(ws.cell(3, 5).value), 5), (clean_text(ws.cell(3, 7).value), 7)]
    day_ranges = [(0, 180), (181, 700)]

    for row in range(5, ws.max_row + 1):
        model = clean_text(ws.cell(row, 2).value)
        if not model:
            continue

        oem_months = int(ws.cell(row, 3).value)
        oem_kms = parse_kms(ws.cell(row, 4).value)
        for plan_name, start_col in plans:
            record = build_record(
                "Audi",
                model,
                model,
                "N/A",
                "N/A",
                oem_months,
                oem_kms,
                plan_name,
                None,
                day_ranges,
                [ws.cell(row, start_col).value, ws.cell(row, start_col + 1).value],
            )
            if record:
                yield record


def parse_bmw(workbook):
    ws = workbook["BMW"]
    day_ranges = [(0, 45), (46, 730)]

    for row in range(4, 37):
        brand = clean_text(ws.cell(row, 1).value)
        model = clean_text(ws.cell(row, 2).value)
        if not brand or not model:
            continue

        fuel = norm_fuel(ws.cell(row, 3).value)
        for plan_name, start_col in [
            ("BMW 3rd Year BRI / UNL Kms", 5),
            ("BMW 4th Year BRI / UNL Kms", 7),
            ("BMW 5th Year BRI / UNL Kms", 9),
        ]:
            record = build_record(
                brand,
                model,
                model,
                fuel,
                "N/A",
                24,
                None,
                plan_name,
                None,
                day_ranges,
                [ws.cell(row, start_col).value, ws.cell(row, start_col + 1).value],
            )
            if record:
                yield record

    for row in range(42, 48):
        brand = clean_text(ws.cell(row, 1).value)
        model = clean_text(ws.cell(row, 2).value)
        if not brand or not model:
            continue

        fuel = norm_fuel(ws.cell(row, 3).value)
        for plan_name, start_col in [
            ("12 Months/UNL Kms", 5),
            ("24 Months/UNL Kms", 7),
            ("36 Months/UNL Kms", 9),
        ]:
            record = build_record(
                brand,
                model,
                model,
                fuel,
                "N/A",
                24,
                None,
                plan_name,
                None,
                day_ranges,
                [ws.cell(row, start_col).value, ws.cell(row, start_col + 1).value],
            )
            if record:
                yield record


def parse_jlr(workbook):
    ws = workbook["JLR"]
    plans = [
        (clean_text(ws.cell(1, 5).value), 5),
        (clean_text(ws.cell(1, 9).value), 9),
        (clean_text(ws.cell(1, 13).value), 13),
    ]
    day_ranges = [(0, 30), (31, 365), (366, 730), (731, 1095)]

    for row in range(3, ws.max_row + 1):
        brand = clean_text(ws.cell(row, 1).value)
        model = clean_text(ws.cell(row, 2).value)
        variant = clean_text(ws.cell(row, 3).value)
        if not brand or not model or not variant:
            continue

        fuel = norm_fuel(ws.cell(row, 4).value)
        for plan_name, start_col in plans:
            record = build_record(
                brand,
                model,
                variant,
                fuel,
                "N/A",
                36,
                None,
                plan_name,
                parse_kms(plan_name),
                day_ranges,
                [ws.cell(row, start_col + i).value for i in range(4)],
            )
            if record:
                yield record


PARSERS = [
    ("Kia", parse_kia),
    ("Tata", parse_tata),
    ("Hyundai", parse_hyundai),
    ("Renault", parse_renault),
    ("Maruti", parse_maruti),
    ("Mahindra", parse_mahindra),
    ("Citroen", parse_citroen),
    ("Toyota", parse_toyota),
    ("MG", parse_mg),
    ("Volkswagen", parse_vw),
    ("Skoda", parse_skoda),
    ("Jeep", parse_jeep),
    ("Honda", parse_honda),
    ("Mercedes-Benz", parse_mbi),
    ("Audi", parse_audi),
    ("BMW/MINI", parse_bmw),
    ("JLR", parse_jlr),
]


def response_data(resp):
    if resp is None:
        return []
    if hasattr(resp, "data"):
        return resp.data or []
    if isinstance(resp, dict):
        return resp.get("data", []) or []
    return []


class DryRunSink:
    def __init__(self):
        self.brand_cache = {}
        self.model_cache = {}
        self.variant_cache = {}
        self.plan_cache = {}
        self.tier_cache = {}
        self.stats = Counter()

    def _fake_id(self):
        return str(uuid.uuid4())

    def get_or_create_brand(self, name):
        key = clean_text(name)
        if key not in self.brand_cache:
            self.brand_cache[key] = self._fake_id()
            self.stats["brands"] += 1
        return self.brand_cache[key]

    def get_or_create_model(self, brand_id, name):
        key = (brand_id, clean_text(name))
        if key not in self.model_cache:
            self.model_cache[key] = self._fake_id()
            self.stats["models"] += 1
        return self.model_cache[key]

    def get_or_create_variant(self, model_id, name, fuel, transmission, oem_warranty_months, oem_warranty_kms):
        key = (model_id, clean_text(name))
        if key not in self.variant_cache:
            self.variant_cache[key] = self._fake_id()
            self.stats["variants"] += 1
        return self.variant_cache[key]

    def get_or_create_plan(self, variant_id, plan_name, duration_months, max_kms):
        key = (variant_id, clean_text(plan_name), duration_months, max_kms)
        if key not in self.plan_cache:
            self.plan_cache[key] = self._fake_id()
            self.stats["plans"] += 1
        return self.plan_cache[key]

    def upsert_tier(self, plan_id, min_days, max_days, price_inr, is_active=True):
        key = (plan_id, min_days, max_days)
        if key not in self.tier_cache:
            self.tier_cache[key] = self._fake_id()
            self.stats["tiers"] += 1
        return self.tier_cache[key]


class SupabaseSink:
    def __init__(self, url, service_key):
        if create_client is None:
            raise RuntimeError("supabase package is not installed. Run: pip install supabase python-dotenv openpyxl")
        self.client = create_client(url, service_key)
        self.brand_cache = {}
        self.model_cache = {}
        self.variant_cache = {}
        self.plan_cache = {}

    def _insert_and_get_id(self, table_name, payload, lookup_filters):
        existing = self.client.table(table_name).select("id")
        for key, value in lookup_filters.items():
            if value is None:
                existing = existing.is_(key, "null")
            else:
                existing = existing.eq(key, value)
        existing_resp = existing.limit(1).execute()
        existing_rows = response_data(existing_resp)
        if existing_rows:
            return existing_rows[0]["id"]

        insert_resp = self.client.table(table_name).insert(payload).execute()
        insert_rows = response_data(insert_resp)
        if insert_rows and "id" in insert_rows[0]:
            return insert_rows[0]["id"]

        reread = self.client.table(table_name).select("id")
        for key, value in lookup_filters.items():
            if value is None:
                reread = reread.is_(key, "null")
            else:
                reread = reread.eq(key, value)
        reread_resp = reread.limit(1).execute()
        reread_rows = response_data(reread_resp)
        if reread_rows:
            return reread_rows[0]["id"]

        raise RuntimeError(f"Could not resolve inserted id for table '{table_name}'")

    def get_or_create_brand(self, name):
        key = clean_text(name)
        if key in self.brand_cache:
            return self.brand_cache[key]
        payload = {"name": key}
        brand_id = self._insert_and_get_id("brands", payload, {"name": key})
        self.brand_cache[key] = brand_id
        return brand_id

    def get_or_create_model(self, brand_id, name):
        key = (brand_id, clean_text(name))
        if key in self.model_cache:
            return self.model_cache[key]
        payload = {"brand_id": brand_id, "name": clean_text(name)}
        model_id = self._insert_and_get_id("models", payload, {"brand_id": brand_id, "name": clean_text(name)})
        self.model_cache[key] = model_id
        return model_id

    def get_or_create_variant(self, model_id, name, fuel, transmission, oem_warranty_months, oem_warranty_kms):
        variant_name = clean_text(name)
        key = (model_id, variant_name)
        if key in self.variant_cache:
            return self.variant_cache[key]

        existing = (
            self.client.table("variants")
            .select("id")
            .eq("model_id", model_id)
            .eq("name", variant_name)
            .limit(1)
            .execute()
        )
        rows = response_data(existing)
        if rows:
            variant_id = rows[0]["id"]
            self.variant_cache[key] = variant_id
            return variant_id

        payload = {
            "model_id": model_id,
            "name": variant_name,
            "fuel": clean_text(fuel),
            "transmission": clean_text(transmission),
            "oem_warranty_months": oem_warranty_months,
            "oem_warranty_kms": oem_warranty_kms,
        }
        lookup = {
            "model_id": model_id,
            "name": variant_name,
        }
        variant_id = self._insert_and_get_id("variants", payload, lookup)
        self.variant_cache[key] = variant_id
        return variant_id

    def get_or_create_plan(self, variant_id, plan_name, duration_months, max_kms):
        key = (variant_id, clean_text(plan_name), duration_months, max_kms)
        if key in self.plan_cache:
            return self.plan_cache[key]

        payload = {
            "variant_id": variant_id,
            "plan_name": clean_text(plan_name),
            "plan_code": None,
            "duration_months": duration_months,
            "max_kms": max_kms,
        }
        lookup = {
            "variant_id": variant_id,
            "plan_name": clean_text(plan_name),
            "duration_months": duration_months,
            "max_kms": max_kms,
        }
        plan_id = self._insert_and_get_id("plans", payload, lookup)
        self.plan_cache[key] = plan_id
        return plan_id

    def upsert_tier(self, plan_id, min_days, max_days, price_inr, is_active=True):
        existing = (
            self.client.table("tiers")
            .select("id")
            .eq("plan_id", plan_id)
            .eq("min_days", min_days)
            .eq("max_days", max_days)
            .limit(1)
            .execute()
        )
        rows = response_data(existing)
        payload = {
            "plan_id": plan_id,
            "min_days": min_days,
            "max_days": max_days,
            "price_inr": price_inr,
            "is_active": is_active,
        }
        if rows:
            tier_id = rows[0]["id"]
            self.client.table("tiers").update(payload).eq("id", tier_id).execute()
            return tier_id

        inserted = self.client.table("tiers").insert(payload).execute()
        data = response_data(inserted)
        if data and "id" in data[0]:
            return data[0]["id"]

        reread = (
            self.client.table("tiers")
            .select("id")
            .eq("plan_id", plan_id)
            .eq("min_days", min_days)
            .eq("max_days", max_days)
            .limit(1)
            .execute()
        )
        reread_rows = response_data(reread)
        if reread_rows:
            return reread_rows[0]["id"]
        raise RuntimeError("Could not insert tier")


def persist_record(sink, record):
    brand_id = sink.get_or_create_brand(record["brand"])
    model_id = sink.get_or_create_model(brand_id, record["model"])
    variant_id = sink.get_or_create_variant(
        model_id,
        record["variant"],
        record["fuel"],
        record["transmission"],
        record["oem_warranty_months"],
        record["oem_warranty_kms"],
    )
    plan_id = sink.get_or_create_plan(
        variant_id,
        record["plan_name"],
        record["duration_months"],
        record["max_kms"],
    )
    for tier in record["tiers"]:
        sink.upsert_tier(
            plan_id,
            tier["min_days"],
            tier["max_days"],
            tier["price_inr"],
            tier["is_active"],
        )


def run_parser(workbook, label, parser_func, step_index=None, step_total=None):
    step_prefix = f"[{step_index}/{step_total}] " if step_index is not None and step_total is not None else ""
    log_info(f"{step_prefix}Parsing {label}")
    records = list(parser_func(workbook))
    log_ok(f"{step_prefix}{label}: parsed {len(records)} plan rows")
    return records


def main():
    parser = argparse.ArgumentParser(description="Ingest Baja Motor EW pricelist into Supabase")
    parser.add_argument("--excel", default=os.getenv("EXCEL_PATH", DEFAULT_EXCEL_PATH), help="Path to the Excel workbook")
    parser.add_argument("--dry-run", action="store_true", help="Parse workbook and print counts without writing to Supabase")
    parser.add_argument("--continue-on-error", action="store_true", help="Skip a failed parser step and continue with the remaining brands")
    args = parser.parse_args()

    load_dotenv()

    if not os.path.exists(args.excel):
        raise FileNotFoundError(f"Excel file not found: {args.excel}")

    log_info(f"Loading workbook: {args.excel}")
    workbook = load_workbook(args.excel, data_only=False)

    all_records = []
    parse_failures = []
    total_parsers = len(PARSERS)

    log_info(f"Starting parse phase for {total_parsers} sheet groups")
    for step_index, (label, parser_func) in enumerate(PARSERS, start=1):
        render_progress("Parsing sheets", step_index - 1, total_parsers)
        try:
            parsed = run_parser(workbook, label, parser_func, step_index, total_parsers)
            all_records.extend(parsed)
        except Exception as exc:
            parse_failures.append((label, str(exc)))
            log_error(f"[{step_index}/{total_parsers}] {label} parser failed: {exc}")
            if not args.continue_on_error:
                finish_progress()
                raise
        finally:
            render_progress("Parsing sheets", step_index, total_parsers)
            finish_progress()

    log_info(f"Parsed {len(all_records)} plan rows total")

    if parse_failures:
        log_warn("Parser failures encountered:")
        for label, err in parse_failures:
            log_warn(f" - {label}: {err}")

    if args.dry_run:
        sink = DryRunSink()
        total_records = len(all_records)
        log_info("Starting dry-run persistence simulation")
        for idx, record in enumerate(all_records, start=1):
            persist_record(sink, record)
            if idx == 1 or idx == total_records or idx % 25 == 0:
                render_progress("Dry run", idx, total_records)
        if total_records:
            finish_progress()
        log_ok(
            "Dry run summary -> "
            f"brands={sink.stats['brands']}, "
            f"models={sink.stats['models']}, "
            f"variants={sink.stats['variants']}, "
            f"plans={sink.stats['plans']}, "
            f"tiers={sink.stats['tiers']}"
        )
        return

    supabase_url = os.getenv("SUPABASE_URL")
    supabase_service_key = os.getenv("SUPABASE_SERVICE_KEY")
    if not supabase_url or not supabase_service_key:
        raise RuntimeError("SUPABASE_URL and SUPABASE_SERVICE_KEY must be set in .env")

    sink = SupabaseSink(supabase_url, supabase_service_key)
    total_records = len(all_records)
    log_info("Starting Supabase write phase")
    for idx, record in enumerate(all_records, start=1):
        for attempt in range(5):
            try:
                persist_record(sink, record)
                break
            except Exception as exc:
                if attempt < 4:
                    wait = 2 ** attempt
                    log_warn(f"Record {idx} attempt {attempt+1} failed ({exc}), retrying in {wait}s...")
                    time.sleep(wait)
                    # Recreate client on connection errors to get a fresh HTTP connection
                    sink = SupabaseSink(supabase_url, supabase_service_key)
                else:
                    raise
        if idx == 1 or idx == total_records or idx % 25 == 0:
            render_progress("Writing to Supabase", idx, total_records)
    if total_records:
        finish_progress()

    log_ok(f"Ingestion completed. Persisted {len(all_records)} plan rows.")


if __name__ == "__main__":
    main()
